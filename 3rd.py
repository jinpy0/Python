"""
    JSON
    [{"홍길동1": 21}, {"홍길동2": 22}, {"홍길동3": 23}]
    CSV
    [{'name': '홍길동1', 'age': 21}, {'name': '홍길동2', 'age': 22}, {'name': '홍길동3', 'age': 23}]

    name,  age
    "홍길동1", 21
    "홍길동2", 22
"""
# CSV : Writer
import openpyxl.workbook

def test1():
    import csv
    
    item_list = []
    for i in range(1, 5):
        dic = {}
        dic["name"] = f"홍길동{i}"
        dic["age"] = 20+i
        item_list.append(dic)
    
    header = ["name", "age"]
    # data = [[],[]]
    with open("test.csv", "w", encoding="utf-8", newline="")as f:
        writer = csv.writer(f,quotechar="'", quoting = csv.QUOTE_NONNUMERIC)
        writer.writerow(header)
        writer.writerow()
        for item in item_list:
            # writer.writerow(item.values())
            writer.writerow([item['name'], item['age']])
            # print(item, "==>", [item['name'], item['age']]) # 딕셔너리에서 리스트로 변환


# test1()
# [{"name" : "홍길동1" , "age" : 21} , {"name" : "홍길동2" , "age" : 22} , ...]
# ************************************************
# CSV : Reader
# ************************************************

def test2():
    import csv
    item_list = []

    with open("test.csv", "r", encoding="utf8") as f:
        reader = csv.reader(f, quotechar="'", quoting=csv.QUOTE_MINIMAL)
        count = 0
        header = []
        for item in reader:
            if count == 0:
                header = item
            else:
                print(item)
                dic = {}
                dic[header[0]] = item[0]  # dic["name"] = "홍길동1"
                dic[header[1]] = item[1]  # dic["age"] = 21
                item_list.append(dic)
            count += 1
    print(item_list)

# def test2():
#     import csv
#     item_list = []
#     with open("test.csv", "r", encoding="utf-8") as f: # 파일 객체를 f에 저장
#         reader = csv.reader(f, quotechar = "'", quoting = csv.QUOTE_NONNUMERIC)
#         count = 1
#         header = []
#         for item in reader: # reader 객체를 통해 csv 파일을 한 줄씩 읽어오고 item 변수에 저장
#             if count == 1:
#                 header = item
#             else:
#                 dic = {}
#                 dic[header[0]] = item[0] # dic["name"] = "홍길동1"
#                 dic[header[1]] = item[1] # dic["age"] = 21
#                 item_list.append(dic)
#             count += 1
#     print(item_list)
        
# test2()

def test3():
    import json
    
    item_list = []
    for i in range(1,6):
        dic = {}
        dic["name"] = f"홍길동{i}"
        dic["age"] = 20+i
        item_list.append(dic)
    print(item_list)
    with open("test.json", "w", encoding="utf-8")as f:
        json.dump(item_list, f, ensure_ascii = False)
    
test3()

def test4():
    import json
    
    item_list = []
    with open("test.json", "r", encoding= "utf-8") as f:
        item_list = json.load(f)
    # print(item_list)    
    
    for item in item_list:
        for k,v in item.items(): # 언패킹
            print(k, ":", v, end = "  ")
        print("")

# test4()

# Excel : Write

def test5():
    import openpyxl
    
    # Excel 만들기
    wb = openpyxl.Workbook()
    
    # Sheet 만들기
    ws = wb.create_sheet("회원정보")
    
    # 데이터 저장하기
    ws["A1"] = "name"
    ws["B1"] = "age"
    ws["A2"] = "홍길동1"
    ws["B2"] = 21
    ws["A3"] = "홍길동2"
    ws["B3"] = 22
    
    # 파일 저장하기
    wb.save("test.xlsx")
    
# test5()

def test5_2():
    import openpyxl

    item_list = []
    for i in range(1, 6):
        dic = {}
        dic["name"] = f"홍길동{i}"
        dic["age"] = 20 + i
        item_list.append(dic)

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("회원정보2")
    ws['A1'] = "name"
    ws['B1'] = "age"
    row = 2
    for item in item_list:
        # print(item)
        # for k,v in item.items():
        ws[f'A{row}'] = item["name"] # A1, A2 ... 에 접근
        ws[f'B{row}'] = item["age"] # B1, B2 ... 에 접근
        row += 1

    wb.save("test.xlsx")

# test5_2()

# Excel : Read
def test6():
    import openpyxl

    item_list = []

    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['회원정보2']

    # ws['A1'].value
    header = [ws['A1'].value, ws['B1'].value]
    for i in range(2, 7):
        dic = {}
        # dic['name'] = ws[f'A[i]'].value
        dic[header[0]] = ws[f'A{i}'].value
        # dic['age'] = ws[f'B[i]'].value
        dic[header[1]] = ws[f'B{i}'].value
        item_list.append(dic)
    print(item_list)

# test6()

# n명의 이름, 국어, 영어, 수학 점수를 파일에 저장하기
# JSON, CSV, Excel
# 시험 문제 공부!!!!
def test7():
    import json
    import openpyxl
    import csv

    item_list = []
    while True:
        st = input("이름, 국어, 영어, 수학 [quit : 종료]")
        if (st.lower() == "quit"):
            break
        score = st.split()
        dic = {}
        dic["name"] = score[0]
        dic["korean"] = int(score[1])
        dic["english"] = int(score[2])
        dic["math"] = int(score[3])
        item_list.append(dic)

    header = ["name", "korean", "english", "math"]
    with open("quiz.json", "w", encoding="utf-8") as f:
        json.dump(item_list, f, ensure_ascii = False)

    with open("quiz.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, quotechar="'", quoting = csv.QUOTE_MINIMAL)
        writer.writerow(header)
        for item in item_list:
            writer.writerow([item["name"], item['korean'], item['english'], item['math']])

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("시험점수")
    ws['A1'] = "name"
    ws['B1'] = "korean"
    ws['C1'] = "english"
    ws['D1'] = "math"
    row = 2
    for item in item_list:
        ws[f'A{row}'] = item["name"]
        ws[f'B{row}'] = item["korean"]
        ws[f'C{row}'] = item["english"]
        ws[f'D{row}'] = item["math"]
        row += 1

    wb.save("quiz.xlsx")
    
# test7()