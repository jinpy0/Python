from bs4 import BeautifulSoup
from urllib.request import urlopen
import csv
import openpyxl
import json

url = "https://www.op.gg/champions?position=top"
html = urlopen(url)
obj_list = BeautifulSoup(html,"html.parser")

tr_list = obj_list.select("#content-container > div.css-67h8o1.esk32cx0 > main > div > table > tbody > tr")

# print(len(tr_list))
#content-container > div.css-67h8o1.esk32cx0 > main > div > table > tbody > tr:nth-child(1)

champion_list = []
champ_dic_list = []
for index, item in enumerate(tr_list):
    dic = {}
    ranking = item.select_one("td.css-v6xa4a.ez7snl14 > span:nth-child(1)").text 
    name = item.select_one("td.css-1hw6gn9.ez7snl12 > a > strong").text
    tier = item.select_one("td.css-1qly9n1.ez7snl15").text
    # win = item.select_one("td:nth-child(5)").text  왜 안됨?
    list = [ranking, name, tier]
    champion_list.append(list)
    dic["랭킹"] = ranking
    dic["이름"] = name
    dic["티어"] = tier
    champ_dic_list.append(dic)
    
header = ["랭킹", "이름", "티어"]    
with open("champion.csv", "w", encoding= "utf-8", newline="")as f:
    writer = csv.writer(f, quotechar = '"', quoting = csv.QUOTE_NONNUMERIC)
    writer.writerow(header)
    for item in champion_list:
        writer.writerow(item)
    
with open("champion.json", "w", encoding = "utf-8")as f:
    json.dump(champ_dic_list, f, ensure_ascii=False)

wb = openpyxl.Workbook()
ws = wb.create_sheet("챔피언 티어표")
ws["A1"] = "진표"
ws["B1"] = "이름"
ws["C1"] = "티어"
start_index = 2
end_index = len(champion_list) + start_index

# index = 0
for i in range(start_index, end_index):
    champion = champion_list[i - start_index]
    ws[f"A{i}"] = champion[0]
    ws[f"B{i}"] = champion[1]
    ws[f"C{i}"] = champion[2]
    # index +=1
    
    
wb.save("champion.xlsx")



# champion_list, champ_dic_list

with open("champ01.csv", "w", encoding = "utf-8", newline = "")as f:
    writer = csv.writer(f, quotechar = '"', quoting = csv.QUOTE_NONNUMERIC)
    writer.writerow(header)
    for i in champion_list:
        writer.writerow(i)
        
        
with open("champ01.json", "w", encoding = "utf-8")as f:
    json.dump(champ_dic_list, f, ensure_ascii = False)
    
wb = openpyxl.Workbook()
ws = wb.create_sheet("챔피언리스트01")
ws["A1"] = "랭킹"
ws["B1"] = "이름"
ws["C1"] = "티어"

st_index = 2
end_index = len(champion_list) + st_index

for i in range(st_index, end_index):
    i_list = champion_list[i-st_index]
    ws[f"A{i}"] = i_list[0]
    ws[f"B{i}"] = i_list[1]
    ws[f"C{i}"] = i_list[2]
    
wb.save("champ01.xlsx")