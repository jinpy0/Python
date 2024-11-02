import json
import csv
import openpyxl
from bs4 import BeautifulSoup
from urllib.request import urlopen

import openpyxl.workbook

url = "https://nenechicken.com/17_new/sub_menu01.asp?SSUBID=1"
html = urlopen(url)
list = BeautifulSoup(html.read(), "html.parser")

div_list = list.select("#menu > div > div > div > div > div")
#menu > div > div > div > div > div:nth-child(1)

# print(len(div_list))
# 길이 확인

menu_list = []

for index, item in enumerate(div_list):
       
    name = item.select_one("div.neneMenu_order > div.neneMenu_order_l > p.neneMenu_bold").text
    #menu > div > div > div > div > div:nth-child(1) > div > div.neneMenu_order > div.neneMenu_order_l > p.neneMenu_bold
    
    script = item.select_one("div.neneMenu_order > div.neneMenu_order_l > p.neneMenu_light").text
    #menu > div > div > div > div > div:nth-child(1) > div > div.neneMenu_order > div.neneMenu_order_l > p.neneMenu_light

    price = item.select_one("div.neneMenu_price > div > span.neneMenu_menuPriceReal").text
    #menu > div > div > div > div > div:nth-child(1) > div > div.neneMenu_price > div > span.neneMenu_menuPriceReal
    
    obj_list = [name, script, price]
    menu_list.append(obj_list)
    
# print(menu_list)

# 
# CSV 파일 저장
# 

header = ["상품명", "설명", "가격"]

with open("chicken.csv", "w", encoding="utf-8", newline="") as f:
    # f.write(header)
    writer = csv.writer(f, quotechar='"', quoting=csv.QUOTE_NONNUMERIC,)
    # writer(header)
    writer.writerow(header)
    for item in menu_list:
        # f.write(item)    
        # writer(item)
        writer.writerow(item)
    
#
# json 파일 저장
# 

# 딕셔너리 리스트 생성
dic_list = []
for item in menu_list:
    dic = {}
    dic["상품명"] = item[0]
    dic["설명"] = item[1]
    dic["가격"] = item[2]
    
    dic_list.append(dic)
    
with open("chicken.json", "w", encoding="utf-8") as f:
    # for item in dic_list:
        # json.dump(item)
    json.dump(dic_list, f, ensure_ascii=False)


#
# Excel 파일 저장
# 

wb = openpyxl.Workbook()
ws = wb.create_sheet("제품 정보")

#
# 딕셔너리 형태로 받아와 저장하기
#

ws["A1"] = "제품명"
ws["B1"] = "설명"
ws["C1"] = "가격"

start_index = 2
end_index = len(dic_list) + start_index

# for item in dic_list:
for i in range(start_index, end_index):
    # ws[f"A{i}"] = dic_list[0]
    a_list = dic_list[i-start_index]
    ws[f"A{i}"] = a_list["상품명"]
    ws[f"B{i}"] = a_list["설명"]
    ws[f"C{i}"] = a_list["가격"]

#
# 리스트 형태로 받아와 저장하기
#

ws1 = wb.create_sheet("제품정보 1")
ws1["A1"] = "제품명"
ws1["B1"] = "설명"
ws1["C1"] = "가격"

start_index = 2
end_index = len(menu_list) + start_index

for i in range(start_index, end_index):
    b_list = menu_list[i-start_index]
    ws1[f"A{i}"] = b_list[0]
    ws1[f"B{i}"] = b_list[1]
    ws1[f"C{i}"] = b_list[2]
    
# ws.
wb.save("chicken.xlsx")

# menu_List , dic_list 저장하기 ( 제품명, 설명, 가격 )

# CSV 저장

header = ["상품명", "설명", "가격"]
with open("chicken02.csv", "w", encoding = "utf-8", newline="")as f:
    # writer = csv.writer()
    writer = csv.writer(f, quotechar='"', quoting = csv.QUOTE_NONNUMERIC)
    writer.writerow(header)
    for item in menu_list:
        writer.writerow(item)
        

# json 저장

with open("chicken02.json", "w", encoding = "utf-8")as f:
    json.dump(dic_list, f, ensure_ascii=False)


# Excel 저장
wb02 = openpyxl.Workbook()
ws02 = wb02.create_sheet("chicken")
ws02["A1"] = "제품명"
ws02["B1"] = "설명"
ws02["C1"] = "가격"
s_index = 2
e_index = len(menu_list) - s_index
for i in range(s_index, e_index):
    item_list = menu_list[i]
    ws02[f"A{i}"] = item_list[0]
    ws02[f"B{i}"] = item_list[1]
    ws02[f"C{i}"] = item_list[2]
    
    
wb.save("chicken02.xlsx")






